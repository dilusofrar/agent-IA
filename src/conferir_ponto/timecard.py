from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
import csv
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PERIOD_PATTERN = re.compile(
    r"In[ií]cio Ponto:\s*(?P<start>\d{2}/\d{2}/\d{4}).*?Fim Ponto:\s*(?P<end>\d{2}/\d{2}/\d{4})",
    re.IGNORECASE | re.DOTALL,
)
EMPLOYEE_PATTERN = re.compile(
    r"Matr[ií]cula:\s*\d+\s*-\s*\d+\s+(?P<name>.+)",
    re.IGNORECASE,
)
DAY_HEADER_PATTERN = re.compile(
    r"^(?P<day>\d{2})\s+(Seg|Ter|Qua|Qui|Sex|Sab|Dom)$",
    re.IGNORECASE,
)
PUNCH_MARK_PATTERN = re.compile(r"(?P<time>\d{2}:\d{2})\s+(?:[odip])\b", re.IGNORECASE)
JOURNEY_PATTERN = re.compile(
    r"Jornada:\s*\d+\s*-\s*(?P<start>\d{2}:\d{2})\s+(?P<lunch_start>\d{2}:\d{2})\s+"
    r"(?P<lunch_end>\d{2}:\d{2})\s+(?P<end>\d{2}:\d{2})",
    re.IGNORECASE,
)
TURN_PATTERN = re.compile(r"(?P<line>\d{2}:\d{2}-\d{2}:\d{2}_.+)", re.IGNORECASE)

OFFICIAL_START = time(7, 45)
OFFICIAL_LUNCH_START = time(12, 0)
OFFICIAL_LUNCH_END = time(13, 0)
OFFICIAL_END = time(17, 0)

DEFAULT_WORKING_WEEKDAYS = frozenset({0, 1, 2, 3, 4})

STATUS_LABELS = {
    "TB": "Trabalhado",
    "FE": "Feriado",
    "CO": "Compensacao",
    "RE": "Repouso",
    "NA": "Nao identificado",
}


@dataclass(frozen=True)
class RawPunchDay:
    work_date: date
    status_code: str
    entries: tuple[str, ...]
    block_lines: tuple[str, ...]


@dataclass(frozen=True)
class DayMetrics:
    work_date: date
    weekday_label: str
    status_code: str
    status_label: str
    first_entry: str | None
    last_exit: str | None
    worked_minutes: int
    expected_minutes: int
    balance_minutes: int
    overtime_before_lunch_minutes: int
    overtime_after_lunch_minutes: int
    payable_overtime_minutes: int
    late_minutes: int
    early_leave_minutes: int
    ignored: bool
    included_in_totals: bool
    holiday_name: str | None
    ignored_reason: str | None = None
    issues: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class WorkSchedule:
    start: time
    lunch_start: time
    lunch_end: time
    end: time
    working_weekdays: frozenset[int] = DEFAULT_WORKING_WEEKDAYS
    source: str | None = None

    @property
    def expected_minutes(self) -> int:
        return int(
            (
                datetime.combine(date.today(), self.lunch_start)
                - datetime.combine(date.today(), self.start)
                + datetime.combine(date.today(), self.end)
                - datetime.combine(date.today(), self.lunch_end)
            ).total_seconds()
            // 60
        )


@dataclass(frozen=True)
class TimeCardAnalysis:
    employee_name: str | None
    period_start: date
    period_end: date
    processed_at: datetime
    schedule: WorkSchedule
    days: list[DayMetrics]

    @property
    def included_days(self) -> list[DayMetrics]:
        return [day for day in self.days if day.included_in_totals]

    @property
    def issues(self) -> list[DayMetrics]:
        return [day for day in self.days if day.issues]


def read_pdf_text(pdf_path: Path) -> str:
    try:
        import fitz
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "PyMuPDF nao esta instalado. Instale as dependencias do projeto."
        ) from exc

    with fitz.open(pdf_path) as document:
        return "\n".join(page.get_text() for page in document)


def read_pdf_bytes_text(content: bytes) -> str:
    try:
        import fitz
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "PyMuPDF nao esta instalado. Instale as dependencias do projeto."
        ) from exc

    with fitz.open(stream=content, filetype="pdf") as document:
        return "\n".join(page.get_text() for page in document)


def parse_timecard_pdf(pdf_path: Path) -> TimeCardAnalysis:
    return parse_timecard_text(read_pdf_text(pdf_path))


def parse_timecard_bytes(content: bytes) -> TimeCardAnalysis:
    return parse_timecard_text(read_pdf_bytes_text(content))


def parse_timecard_text(text: str) -> TimeCardAnalysis:
    period_start, period_end = extract_period_dates(text)
    employee_name = extract_employee_name(text)
    schedule = extract_work_schedule(text)
    raw_days = parse_raw_days(text, period_start, period_end)
    day_map = {raw.work_date: raw for raw in raw_days}

    return TimeCardAnalysis(
        employee_name=employee_name,
        period_start=period_start,
        period_end=period_end,
        processed_at=datetime.now(),
        schedule=schedule,
        days=[
            calculate_day_metrics(current_date, day_map.get(current_date), schedule)
            for current_date in daterange(period_start, period_end)
        ],
    )


def extract_period_dates(text: str) -> tuple[date, date]:
    match = PERIOD_PATTERN.search(text)
    if not match:
        raise ValueError("Nao foi possivel identificar o periodo de apuracao no PDF.")

    start_date = date.fromisoformat("-".join(reversed(match.group("start").split("/"))))
    end_date = date.fromisoformat("-".join(reversed(match.group("end").split("/"))))
    return start_date, end_date


def extract_employee_name(text: str) -> str | None:
    match = EMPLOYEE_PATTERN.search(text)
    if not match:
        return None
    return " ".join(match.group("name").split())


def extract_work_schedule(text: str) -> WorkSchedule:
    working_weekdays = extract_working_weekdays(text)
    source_match = TURN_PATTERN.search(text)
    source = source_match.group("line") if source_match else None

    return WorkSchedule(
        start=OFFICIAL_START,
        lunch_start=OFFICIAL_LUNCH_START,
        lunch_end=OFFICIAL_LUNCH_END,
        end=OFFICIAL_END,
        working_weekdays=working_weekdays,
        source=source,
    )


def extract_working_weekdays(text: str) -> frozenset[int]:
    upper_text = text.upper()
    if "SEG A SEX" in upper_text or "SEG A SE" in upper_text or "SEG À SEX" in upper_text:
        return frozenset({0, 1, 2, 3, 4})
    if "SEG A SAB" in upper_text or "SEG À SAB" in upper_text:
        return frozenset({0, 1, 2, 3, 4, 5})
    if "DOM A DOM" in upper_text:
        return frozenset({0, 1, 2, 3, 4, 5, 6})
    return DEFAULT_WORKING_WEEKDAYS


def parse_raw_days(text: str, period_start: date, period_end: date) -> list[RawPunchDay]:
    lines = [line.strip() for line in text.splitlines()]
    blocks: list[tuple[int, list[str]]] = []
    current_day: int | None = None
    current_lines: list[str] = []

    for line in lines:
        if not line:
            continue

        day_match = DAY_HEADER_PATTERN.match(line)
        if day_match:
            if current_day is not None:
                blocks.append((current_day, current_lines))
            current_day = int(day_match.group("day"))
            current_lines = [line]
            continue

        if current_day is not None:
            current_lines.append(line)

    if current_day is not None:
        blocks.append((current_day, current_lines))

    day_lookup = build_period_day_lookup(period_start, period_end)
    last_date: date | None = None
    raw_days: list[RawPunchDay] = []

    for day_number, block_lines in blocks:
        work_date = resolve_work_date(day_number, day_lookup, last_date)
        if work_date is None:
            continue

        entries = extract_entries_from_block(block_lines)
        raw_days.append(
            RawPunchDay(
                work_date=work_date,
                status_code=detect_status_code(block_lines),
                entries=entries,
                block_lines=tuple(block_lines),
            )
        )
        last_date = work_date

    return raw_days


def detect_status_code(block_lines: list[str]) -> str:
    for line in block_lines:
        normalized = line.strip().upper()
        if normalized in STATUS_LABELS:
            return normalized
    return "NA"


def extract_entries_from_block(block_lines: list[str]) -> tuple[str, ...]:
    best_match: tuple[str, ...] = ()
    for line in block_lines:
        matches = tuple(match.group("time") for match in PUNCH_MARK_PATTERN.finditer(line))
        if len(matches) > len(best_match):
            best_match = matches
    return best_match


def is_bridge_day(raw_day: RawPunchDay) -> bool:
    for line in raw_day.block_lines:
        normalized = line.upper()
        if "PONTE DE FERIADO" in normalized:
            return True
        if normalized in {"PONTE", "DIA PONTE"}:
            return True
    return False


def detect_ignored_reason(raw_day: RawPunchDay) -> str | None:
    block_text = " ".join(raw_day.block_lines).upper()
    if "FERIAS" in block_text:
        return "Ferias"
    if "COMPENSACAO F_DIA" in block_text:
        return "Compensacao de feriado"
    if is_bridge_day(raw_day):
        return "Dia ponte"
    return None


def build_period_day_lookup(start_date: date, end_date: date) -> dict[int, list[date]]:
    lookup: dict[int, list[date]] = {}
    current = start_date
    while current <= end_date:
        lookup.setdefault(current.day, []).append(current)
        current += timedelta(days=1)
    return lookup


def resolve_work_date(
    day_number: int,
    day_lookup: dict[int, list[date]],
    last_date: date | None,
) -> date | None:
    candidates = day_lookup.get(day_number, [])
    if not candidates:
        return None
    if last_date is None:
        return candidates[0]
    for candidate in candidates:
        if candidate >= last_date:
            return candidate
    return candidates[-1]


def calculate_day_metrics(work_date: date, raw_day: RawPunchDay | None, schedule: WorkSchedule) -> DayMetrics:
    holiday_name = get_brazil_holiday_name(work_date)
    non_working_weekday = work_date.weekday() not in schedule.working_weekdays

    if raw_day is None:
        if non_working_weekday or holiday_name:
            return build_ignored_day(work_date, "NA", holiday_name)
        return DayMetrics(
            work_date=work_date,
            weekday_label=weekday_pt(work_date),
            status_code="NA",
            status_label=STATUS_LABELS["NA"],
            first_entry=None,
            last_exit=None,
            worked_minutes=0,
            expected_minutes=schedule.expected_minutes,
            balance_minutes=0,
            overtime_before_lunch_minutes=0,
            overtime_after_lunch_minutes=0,
            payable_overtime_minutes=0,
            late_minutes=0,
            early_leave_minutes=0,
            ignored=False,
            included_in_totals=False,
            holiday_name=holiday_name,
            ignored_reason=None,
            issues=["Dia util sem batidas registradas."],
        )

    status_code = raw_day.status_code
    ignored_reason = detect_ignored_reason(raw_day)
    non_business_day = non_working_weekday or holiday_name or status_code in {"FE", "RE"}
    if ignored_reason:
        return build_ignored_day(work_date, status_code, holiday_name, raw_day, ignored_reason)
    if (non_business_day or status_code == "CO") and not raw_day.entries:
        return build_ignored_day(work_date, status_code, holiday_name, raw_day)

    if len(raw_day.entries) < 2:
        return DayMetrics(
            work_date=work_date,
            weekday_label=weekday_pt(work_date),
            status_code=status_code,
            status_label=STATUS_LABELS.get(status_code, STATUS_LABELS["NA"]),
            first_entry=raw_day.entries[0] if raw_day.entries else None,
            last_exit=raw_day.entries[-1] if raw_day.entries else None,
            worked_minutes=0,
            expected_minutes=schedule.expected_minutes,
            balance_minutes=0,
            overtime_before_lunch_minutes=0,
            overtime_after_lunch_minutes=0,
            payable_overtime_minutes=0,
            late_minutes=0,
            early_leave_minutes=0,
            ignored=False,
            included_in_totals=False,
            holiday_name=holiday_name,
            ignored_reason=None,
            issues=["Quantidade insuficiente de batidas para calcular a jornada."],
        )

    first_entry = raw_day.entries[0]
    last_exit = raw_day.entries[-1]
    punch_datetimes = build_punch_datetimes(work_date, raw_day.entries)
    start_dt = punch_datetimes[0]
    end_dt = punch_datetimes[-1]
    issues: list[str] = []

    if len(punch_datetimes) % 2 != 0:
        issues.append("Quantidade impar de batidas, impossivel fechar os intervalos.")
        return DayMetrics(
            work_date=work_date,
            weekday_label=weekday_pt(work_date),
            status_code=status_code,
            status_label=STATUS_LABELS.get(status_code, STATUS_LABELS["NA"]),
            first_entry=first_entry,
            last_exit=last_exit,
            worked_minutes=0,
            expected_minutes=schedule.expected_minutes,
            balance_minutes=0,
            overtime_before_lunch_minutes=0,
            overtime_after_lunch_minutes=0,
            payable_overtime_minutes=0,
            late_minutes=0,
            early_leave_minutes=0,
            ignored=False,
            included_in_totals=False,
            holiday_name=holiday_name,
            ignored_reason=None,
            issues=issues,
        )

    worked_minutes = calculate_paired_worked_minutes(punch_datetimes)
    if non_business_day or status_code == "CO":
        return build_non_business_workday(
            work_date=work_date,
            status_code=status_code,
            holiday_name=holiday_name,
            first_entry=first_entry,
            last_exit=last_exit,
            worked_minutes=worked_minutes,
            punch_datetimes=punch_datetimes,
            schedule=schedule,
        )

    standard_start_dt = datetime.combine(work_date, schedule.start)
    standard_end_dt = datetime.combine(work_date, schedule.end)
    overtime_before = max(0, minutes_between(start_dt, standard_start_dt))
    overtime_after = max(0, minutes_between(standard_end_dt, end_dt))
    late_minutes = max(0, minutes_between(standard_start_dt, start_dt))
    early_leave_minutes = max(0, minutes_between(end_dt, standard_end_dt))
    balance_minutes = worked_minutes - schedule.expected_minutes

    if start_dt.time() >= schedule.lunch_start and start_dt.time() < schedule.lunch_end:
        issues.append("Entrada registrada durante o intervalo de almoco.")
    if end_dt.time() > schedule.lunch_start and end_dt.time() <= schedule.lunch_end:
        issues.append("Saida registrada dentro do intervalo de almoco.")

    return DayMetrics(
        work_date=work_date,
        weekday_label=weekday_pt(work_date),
        status_code=status_code,
        status_label=STATUS_LABELS.get(status_code, STATUS_LABELS["NA"]),
        first_entry=first_entry,
        last_exit=last_exit,
        worked_minutes=worked_minutes,
        expected_minutes=schedule.expected_minutes,
        balance_minutes=balance_minutes,
        overtime_before_lunch_minutes=overtime_before,
        overtime_after_lunch_minutes=overtime_after,
        payable_overtime_minutes=0,
        late_minutes=late_minutes,
        early_leave_minutes=early_leave_minutes,
        ignored=False,
        included_in_totals=not issues,
        holiday_name=holiday_name,
        ignored_reason=None,
        issues=issues,
    )


def build_ignored_day(
    work_date: date,
    status_code: str,
    holiday_name: str | None,
    raw_day: RawPunchDay | None = None,
    ignored_reason: str | None = None,
) -> DayMetrics:
    first_entry = raw_day.entries[0] if raw_day and raw_day.entries else None
    last_exit = raw_day.entries[-1] if raw_day and raw_day.entries else None
    return DayMetrics(
        work_date=work_date,
        weekday_label=weekday_pt(work_date),
        status_code=status_code,
        status_label=STATUS_LABELS.get(status_code, STATUS_LABELS["NA"]),
        first_entry=first_entry,
        last_exit=last_exit,
        worked_minutes=0,
        expected_minutes=0,
        balance_minutes=0,
        overtime_before_lunch_minutes=0,
        overtime_after_lunch_minutes=0,
        payable_overtime_minutes=0,
        late_minutes=0,
        early_leave_minutes=0,
        ignored=True,
        included_in_totals=False,
        holiday_name=holiday_name,
        ignored_reason=ignored_reason,
        issues=[],
    )


def build_non_business_workday(
    work_date: date,
    status_code: str,
    holiday_name: str | None,
    first_entry: str,
    last_exit: str,
    worked_minutes: int,
    punch_datetimes: list[datetime],
    schedule: WorkSchedule,
) -> DayMetrics:
    morning_minutes, afternoon_minutes = split_paired_minutes_by_lunch(punch_datetimes, schedule)
    return DayMetrics(
        work_date=work_date,
        weekday_label=weekday_pt(work_date),
        status_code=status_code,
        status_label=STATUS_LABELS.get(status_code, STATUS_LABELS["NA"]),
        first_entry=first_entry,
        last_exit=last_exit,
        worked_minutes=worked_minutes,
        expected_minutes=0,
        balance_minutes=0,
        overtime_before_lunch_minutes=morning_minutes,
        overtime_after_lunch_minutes=afternoon_minutes,
        payable_overtime_minutes=worked_minutes,
        late_minutes=0,
        early_leave_minutes=0,
        ignored=False,
        included_in_totals=True,
        holiday_name=holiday_name,
        ignored_reason=None,
        issues=[],
    )


def calculate_worked_minutes(start_dt: datetime, end_dt: datetime) -> int:
    morning, afternoon = split_interval_by_lunch(
        start_dt,
        end_dt,
        lunch_start=start_dt.replace(hour=12, minute=0),
        lunch_end=start_dt.replace(hour=13, minute=0),
    )
    return morning + afternoon


def split_minutes_by_lunch(
    work_date: date,
    first_entry: str,
    last_exit: str,
    schedule: WorkSchedule,
) -> tuple[int, int]:
    return split_interval_by_lunch(
        combine(work_date, first_entry),
        combine(work_date, last_exit),
        lunch_start=datetime.combine(work_date, schedule.lunch_start),
        lunch_end=datetime.combine(work_date, schedule.lunch_end),
    )


def split_paired_minutes_by_lunch(
    punch_datetimes: list[datetime],
    schedule: WorkSchedule,
) -> tuple[int, int]:
    morning = 0
    afternoon = 0
    for start_dt, end_dt in pairwise_datetimes(punch_datetimes):
        current_morning, current_afternoon = split_interval_by_lunch(
            start_dt,
            end_dt,
            lunch_start=datetime.combine(start_dt.date(), schedule.lunch_start),
            lunch_end=datetime.combine(start_dt.date(), schedule.lunch_end),
        )
        morning += current_morning
        afternoon += current_afternoon
    return morning, afternoon


def split_interval_by_lunch(
    start_dt: datetime,
    end_dt: datetime,
    lunch_start: datetime,
    lunch_end: datetime,
) -> tuple[int, int]:
    day_start = start_dt.replace(hour=0, minute=0)
    day_end = end_dt.replace(hour=23, minute=59)
    morning = overlap_minutes(start_dt, end_dt, day_start, lunch_start)
    afternoon = overlap_minutes(start_dt, end_dt, lunch_end, day_end)
    return morning, afternoon


def overlap_minutes(start_a: datetime, end_a: datetime, start_b: datetime, end_b: datetime) -> int:
    start = max(start_a, start_b)
    end = min(end_a, end_b)
    if end <= start:
        return 0
    return int((end - start).total_seconds() // 60)


def combine(work_date: date, clock: str) -> datetime:
    hours, minutes = clock.split(":")
    return datetime.combine(work_date, time(int(hours), int(minutes)))


def parse_clock(value: str) -> time:
    hours, minutes = value.split(":")
    return time(int(hours), int(minutes))


def build_punch_datetimes(work_date: date, entries: tuple[str, ...]) -> list[datetime]:
    datetimes: list[datetime] = []
    current_day = work_date
    previous_dt: datetime | None = None

    for clock in entries:
        current_dt = combine(current_day, clock)
        if previous_dt is not None and current_dt <= previous_dt:
            current_day += timedelta(days=1)
            current_dt = combine(current_day, clock)
        datetimes.append(current_dt)
        previous_dt = current_dt

    return datetimes


def pairwise_datetimes(punch_datetimes: list[datetime]):
    for index in range(0, len(punch_datetimes), 2):
        if index + 1 >= len(punch_datetimes):
            break
        yield punch_datetimes[index], punch_datetimes[index + 1]


def calculate_paired_worked_minutes(punch_datetimes: list[datetime]) -> int:
    total = 0
    for start_dt, end_dt in pairwise_datetimes(punch_datetimes):
        if end_dt <= start_dt:
            continue
        total += int((end_dt - start_dt).total_seconds() // 60)
    return total


def minutes_between(start_dt: datetime, end_dt: datetime) -> int:
    return int((end_dt - start_dt).total_seconds() // 60)


def daterange(start_date: date, end_date: date):
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def weekday_pt(work_date: date) -> str:
    labels = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
    return labels[work_date.weekday()]


def get_brazil_holiday_name(work_date: date) -> str | None:
    try:
        import holidays

        calendar = holidays.country_holidays("BR", years=work_date.year, language="pt_BR")
        holiday_name = calendar.get(work_date)
        if holiday_name:
            return str(holiday_name)
    except Exception:
        pass

    easter = easter_sunday(work_date.year)
    holidays = {
        date(work_date.year, 1, 1): "Confraternizacao Universal",
        easter - timedelta(days=48): "Carnaval",
        easter - timedelta(days=47): "Carnaval",
        easter - timedelta(days=2): "Paixao de Cristo",
        easter: "Pascoa",
        date(work_date.year, 4, 21): "Tiradentes",
        date(work_date.year, 5, 1): "Dia do Trabalhador",
        easter + timedelta(days=60): "Corpus Christi",
        date(work_date.year, 9, 7): "Independencia do Brasil",
        date(work_date.year, 10, 12): "Nossa Senhora Aparecida",
        date(work_date.year, 11, 2): "Finados",
        date(work_date.year, 11, 15): "Proclamacao da Republica",
        date(work_date.year, 11, 20): "Dia da Consciencia Negra",
        date(work_date.year, 12, 25): "Natal",
    }
    return holidays.get(work_date)


def easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def format_minutes(value: int) -> str:
    sign = "-" if value < 0 else ""
    absolute = abs(value)
    hours, minutes = divmod(absolute, 60)
    return f"{sign}{hours:02d}:{minutes:02d}"


def build_summary_payload(analysis: TimeCardAnalysis) -> dict:
    included = analysis.included_days
    total_worked = sum(day.worked_minutes for day in included)
    total_expected = sum(day.expected_minutes for day in included)
    bank_days = [day for day in included if day.expected_minutes > 0]
    paid_overtime_days = [day for day in included if day.payable_overtime_minutes > 0]
    total_balance = sum(day.balance_minutes for day in bank_days)
    total_overtime_before = sum(day.overtime_before_lunch_minutes for day in bank_days)
    total_overtime_after = sum(day.overtime_after_lunch_minutes for day in bank_days)
    total_paid_overtime = sum(day.payable_overtime_minutes for day in paid_overtime_days)
    total_positive = total_overtime_before + total_overtime_after
    total_negative = sum(day.late_minutes + day.early_leave_minutes for day in bank_days)
    total_late = sum(day.late_minutes for day in bank_days)
    total_early = sum(day.early_leave_minutes for day in bank_days)

    return {
        "employeeName": analysis.employee_name,
        "periodStart": analysis.period_start.isoformat(),
        "periodEnd": analysis.period_end.isoformat(),
        "processedAt": analysis.processed_at.isoformat(timespec="seconds"),
        "schedule": {
            "start": analysis.schedule.start.strftime("%H:%M"),
            "lunchStart": analysis.schedule.lunch_start.strftime("%H:%M"),
            "lunchEnd": analysis.schedule.lunch_end.strftime("%H:%M"),
            "end": analysis.schedule.end.strftime("%H:%M"),
            "workingWeekdays": sorted(analysis.schedule.working_weekdays),
            "source": analysis.schedule.source,
        },
        "summary": {
            "businessDaysProcessed": len(included),
            "ignoredDays": len([day for day in analysis.days if day.ignored]),
            "inconsistencyCount": len(analysis.issues),
            "worked": format_minutes(total_worked),
            "expected": format_minutes(total_expected),
            "balance": format_minutes(total_balance),
            "positiveBank": format_minutes(total_positive),
            "negativeBank": format_minutes(total_negative),
            "compensated": format_minutes(total_negative),
            "paidOvertime": format_minutes(total_paid_overtime),
            "overtimeBeforeLunch": format_minutes(total_overtime_before),
            "overtimeAfterLunch": format_minutes(total_overtime_after),
            "late": format_minutes(total_late),
            "earlyLeave": format_minutes(total_early),
        },
        "days": [
            {
                "date": day.work_date.isoformat(),
                "weekday": day.weekday_label,
                "statusCode": day.status_code,
                "statusLabel": day.status_label,
                "holidayName": day.holiday_name,
                "firstEntry": day.first_entry,
                "lastExit": day.last_exit,
                "worked": format_minutes(day.worked_minutes),
                "expected": format_minutes(day.expected_minutes),
                "balance": format_minutes(day.balance_minutes),
                "overtimeBeforeLunch": format_minutes(day.overtime_before_lunch_minutes),
                "overtimeAfterLunch": format_minutes(day.overtime_after_lunch_minutes),
                "paidOvertime": format_minutes(day.payable_overtime_minutes),
                "late": format_minutes(day.late_minutes),
                "earlyLeave": format_minutes(day.early_leave_minutes),
                "ignored": day.ignored,
                "ignoredReason": day.ignored_reason,
                "includedInTotals": day.included_in_totals,
                "issues": day.issues,
            }
            for day in analysis.days
        ],
    }


def export_analysis_to_xlsx(analysis: TimeCardAnalysis) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    details_sheet = workbook.create_sheet("Dias")

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    issue_fill = PatternFill("solid", fgColor="FDE68A")
    ignored_fill = PatternFill("solid", fgColor="E2E8F0")

    summary_rows = [
        ("Colaborador", analysis.employee_name or "Nao identificado"),
        ("Periodo inicial", analysis.period_start.isoformat()),
        ("Periodo final", analysis.period_end.isoformat()),
        ("Processado em", analysis.processed_at.isoformat(timespec="seconds")),
    ]
    payload = build_summary_payload(analysis)["summary"]
    for label, value in payload.items():
        summary_rows.append((label, value))

    for row_index, (label, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row_index, 1).value = label
        summary_sheet.cell(row_index, 2).value = value

    headers = [
        "Data",
        "Dia",
        "Status",
        "Feriado",
        "Entrada",
        "Saida",
        "Trabalhado",
        "Esperado",
        "Saldo",
        "Extra antes almoco",
        "Extra apos almoco",
        "Hora extra paga",
        "Atraso",
        "Saida antecipada",
        "Ignorado",
        "Inconsistencias",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = details_sheet.cell(1, column_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for row_index, day in enumerate(analysis.days, start=2):
        row = [
            day.work_date.isoformat(),
            day.weekday_label,
            day.status_label,
            day.holiday_name,
            day.first_entry,
            day.last_exit,
            format_minutes(day.worked_minutes),
            format_minutes(day.expected_minutes),
            format_minutes(day.balance_minutes),
            format_minutes(day.overtime_before_lunch_minutes),
            format_minutes(day.overtime_after_lunch_minutes),
            format_minutes(day.payable_overtime_minutes),
            format_minutes(day.late_minutes),
            format_minutes(day.early_leave_minutes),
            "Sim" if day.ignored else "Nao",
            " | ".join(day.issues),
        ]
        for column_index, value in enumerate(row, start=1):
            details_sheet.cell(row_index, column_index).value = value

        if day.issues:
            for column_index in range(1, len(headers) + 1):
                details_sheet.cell(row_index, column_index).fill = issue_fill
        elif day.ignored:
            for column_index in range(1, len(headers) + 1):
                details_sheet.cell(row_index, column_index).fill = ignored_fill

    for sheet in (summary_sheet, details_sheet):
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 28)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_analysis_csv(analysis: TimeCardAnalysis, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "Data",
                "Entrada",
                "Saida",
                "Status",
                "Trabalhado",
                "Saldo",
                "Inconsistencias",
            ],
        )
        writer.writeheader()
        for day in analysis.days:
            writer.writerow(
                {
                    "Data": day.work_date.isoformat(),
                    "Entrada": day.first_entry or "",
                    "Saida": day.last_exit or "",
                    "Status": day.status_label,
                    "Trabalhado": format_minutes(day.worked_minutes),
                    "Saldo": format_minutes(day.balance_minutes),
                    "Inconsistencias": " | ".join(day.issues),
                }
            )
