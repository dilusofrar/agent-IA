from __future__ import annotations

import argparse
import csv
import warnings
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CSV = PROJECT_ROOT / "data" / "outputs" / "DIEGO_LUCAS_SOARES_DE_FREITAS_ARAUJO_extraido.csv"
DEFAULT_TEMPLATE = PROJECT_ROOT / "data" / "templates" / "HORAS.xlsm"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "outputs" / "HORAS_teste_macro.xlsm"

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

PORTUGUESE_MONTHS = {
    1: "jan",
    2: "fev",
    3: "mar",
    4: "abr",
    5: "mai",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "set",
    10: "out",
    11: "nov",
    12: "dez",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Automatiza o teste funcional da planilha HORAS.xlsm usando a mesma regra do macro."
    )
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV, help="CSV extraido do cartao ponto.")
    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="Planilha modelo .xlsm.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Arquivo .xlsm preenchido para validacao.",
    )
    parser.add_argument(
        "--sheet",
        default="Folha1",
        help="Nome da aba onde as colunas E, F e I sao preenchidas.",
    )
    return parser.parse_args()


def load_csv_rows(csv_path: Path) -> list[tuple[date, str, str]]:
    rows: list[tuple[date, str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row in reader:
            work_date = parse_csv_date(row["Data"])
            rows.append((work_date, row["Entrada"], row["Saida"]))
    return rows


def parse_csv_date(value: str) -> date:
    day_text, month_text = value.strip().split("/")
    month = MONTHS[month_text.lower()]
    day = int(day_text)
    return date(2025, month, day)


def extract_period_start(ws) -> date:
    start_value = ws["B10"].value
    if isinstance(start_value, datetime):
        return start_value.date()
    if isinstance(start_value, date):
        return start_value
    raise ValueError("Nao foi possivel identificar a data inicial do periodo na celula B10.")


def row_for_date(period_start: date, current_date: date) -> int:
    delta = (current_date - period_start).days
    return 3 + delta


def fill_worksheet(ws, csv_rows: list[tuple[date, str, str]]) -> list[tuple[str, int, str, str]]:
    period_start = extract_period_start(ws)
    validations: list[tuple[str, int, str, str]] = []

    for work_date, entry, exit_time in csv_rows:
        target_row = row_for_date(period_start, work_date)
        ws.cell(target_row, 6).value = entry
        ws.cell(target_row, 9).value = exit_time
        validations.append((format_date_pt(work_date), target_row, entry, exit_time))

    return validations


def format_date_pt(value: date) -> str:
    return f"{value.day:02d}/{PORTUGUESE_MONTHS[value.month]}"


def main() -> int:
    warnings.filterwarnings(
        "ignore",
        category=UserWarning,
        module="openpyxl.worksheet._reader",
    )
    args = parse_args()
    csv_path = args.csv.resolve()
    template_path = args.template.resolve()
    output_path = args.output.resolve()

    if not csv_path.exists():
        raise SystemExit(f"CSV nao encontrado: {csv_path}")

    if not template_path.exists():
        raise SystemExit(f"Planilha modelo nao encontrada: {template_path}")

    csv_rows = load_csv_rows(csv_path)
    workbook = load_workbook(template_path, keep_vba=True)
    worksheet = workbook[args.sheet]
    validations = fill_worksheet(worksheet, csv_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"Arquivo de teste gerado: {output_path}")
    print(f"Registros aplicados na planilha: {len(validations)}")

    for label in ("16/abr", "17/abr", "05/mai", "15/mai"):
        found = next((item for item in validations if item[0] == label), None)
        if found is None:
            print(f"Validacao: data {label} nao encontrada no CSV.")
            continue
        print(
            f"Validacao: {found[0]} | linha {found[1]} | entrada={found[2]} | saida={found[3]}"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
